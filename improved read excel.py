import csv
import pandas as pd
import os
import numpy as np
import openpyxl

def read_file(file_path):
    _, file_extension = os.path.splitext(file_path)
    if file_extension == ".csv":
        with open(file_path, "r") as f:
            reader = csv.DictReader(f)
            rows = [row for row in reader]
    elif file_extension == ".xlsx":
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        rows = []
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            header = [cell.value for cell in sheet[1]]
            sheet_rows = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_dict = {header[i]: value for i, value in enumerate(row)}
                sheet_rows.append(row_dict)
            rows.extend(sheet_rows)
    else:
        raise ValueError("Unsupported file format")
    return rows
